import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

FIRST = ["James", "Mary", "Robert", "Patricia", "John", "Jennifer", "Michael", "Linda",
        "David", "Elizabeth", "William", "Barbara", "Richard", "Susan", "Joseph", "Jessica",
        "Thomas", "Sarah", "Charles", "Karen", "Daniel", "Nancy", "Matthew", "Lisa",
        "Anthony", "Margaret", "Mark", "Betty", "Donald", "Sandra"]

LAST = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis",
        "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez", "Wilson", "Anderson",
        "Thomas", "Taylor", "Moore", "Jackson", "Martin", "Lee", "Perez", "Thompson",
        "White", "Harris", "Sanchez", "Clark", "Ramirez", "Lewis", "Robinson"]

STREETS = ["Main St", "Oak Ave", "Pine Rd", "Maple Dr", "Cedar Ln", "Elm St", "Washington Ave",
           "Lake Dr", "Hillcrest Rd", "Sunset Blvd", "Park Ave", "Church St", "Spring St",
           "River Rd", "Highland Ave", "Meadow Ln", "Ridge Rd", "Broadway", "Jefferson St"]

CITIES_STATES = [
    ("Austin", "TX", "78701"), ("Denver", "CO", "80202"), ("Seattle", "WA", "98101"),
    ("Portland", "OR", "97201"), ("Boston", "MA", "02108"), ("Chicago", "IL", "60601"),
    ("Atlanta", "GA", "30301"), ("Phoenix", "AZ", "85001"), ("Miami", "FL", "33101"),
    ("Nashville", "TN", "37201"), ("Minneapolis", "MN", "55401"), ("Raleigh", "NC", "27601"),
    ("Columbus", "OH", "43215"), ("Indianapolis", "IN", "46201"), ("Charlotte", "NC", "28202"),
    ("San Diego", "CA", "92101"), ("Kansas City", "MO", "64108"), ("Pittsburgh", "PA", "15222"),
    ("Baltimore", "MD", "21201"), ("Las Vegas", "NV", "89101"),
]

wb = Workbook()
ws = wb.active
ws.title = "People"

# Header
ws["A1"] = "Name"
ws["B1"] = "Address"
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill("solid", fgColor="2E5C8A")
for cell in (ws["A1"], ws["B1"]):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")

# Data rows
for row in range(2, 12):
    name = f"{random.choice(FIRST)} {random.choice(LAST)}"
    num = random.randint(100, 9999)
    street = random.choice(STREETS)
    city, state, zipcode = random.choice(CITIES_STATES)
    address = f"{num} {street}, {city}, {state} {zipcode}"
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=address)

# Column widths
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 50
ws.row_dimensions[1].height = 22

wb.save("/Users/travispatterson/Desktop/ClaudeTest1/people.xlsx")
print("Saved people.xlsx")
